
import openpyxl

estoque_file = openpyxl.load_workbook("estoque.xlsx")
lista_produtos = estoque_file['Sheet1']

produtos_por_fornecedor = {}
qtd_por_fornecedor = 0
valor_por_forncedor = 0

print(lista_produtos.iter_cols())

forn_pesquisado = input("Qual empresa você deseja consultar informações? ")
for linha_produtos in range(2, lista_produtos.max_row+1):
    produto = lista_produtos.cell(linha_produtos, 1).value
    estoque = lista_produtos.cell(linha_produtos,2).value
    preco = lista_produtos.cell(linha_produtos,3).value
    fornecedor = str(lista_produtos.cell(linha_produtos, 4).value)
    
    if forn_pesquisado == fornecedor:
        qtd_por_fornecedor += estoque
        valor_por_forncedor += preco
    

print(f'{forn_pesquisado} tem {qtd_por_fornecedor} produtos em estoque, no valor de R$ {valor_por_forncedor:,.2f}.')